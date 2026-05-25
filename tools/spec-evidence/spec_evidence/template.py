"""Generate an EMPTY 'IT確認書' workbook template.

Mirrors the layout of the reference IT test-confirmation book: a cover sheet,
one checklist sheet per category (9 columns, OK/NG dropdown), and an evidence
sheet with numbered screenshot-paste blocks. Cells are left blank for manual
fill-in; a later step can populate them from cases.json + results.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CATEGORY_COLUMNS = [
    "No.", "区分", "テスト項目", "確認手順", "利用データNo.", "期待結果",
    "結果(OK/NG)", "NG時の内容・備考", "エビデンスNo.", "実施日",
]
RESULT_CHOICES = '"OK,NG,-,保留"'
DEFAULT_CATEGORIES = ["画面表示", "ボタン操作", "API連携", "権限・認証"]

DATA_COLUMNS = ["利用データNo.", "区分", "内容", "値・例", "備考"]
DEFAULT_DATA_SETS = [
    ("DT-001", "認証アカウント", "ログイン用モックアカウント", "demo / password1（非永続トークン）", ""),
    ("DT-002", "固定フィクスチャ", "Items一覧/詳細の MSW 固定データ",
     "Coffee(drink/¥350) / Sandwich(food/¥480) / Notebook(other/¥220)", ""),
    ("DT-003", "境界値データ", "gen-cases 生成の境界値",
     "tests/cases/*.cases.json（各値は各行の確認手順に記載）", ""),
    ("DT-004", "環境", "実行環境/モック",
     "DEV=MSW固定（本番ビルドはMSW無し→MSW-in-build/backend要）", ""),
]

_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_TITLE_FONT = Font(bold=True, size=14)
_BOLD = Font(bold=True)
_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _label(ws, cell, text):
    ws[cell] = text
    ws[cell].font = _BOLD


def _build_cover(ws, meta):
    ws.title = "01_表紙"
    ws["B2"] = "IT テスト確認書"
    ws["B2"].font = _TITLE_FONT
    ws["B4"] = f"対象機能: {meta.get('feature', '<記入>')}"
    rows = [
        ("バージョン", meta.get("version", "v1.0")),
        ("発行日", meta.get("issued", "")),
        ("対象URL", meta.get("url", "")),
        ("テスト環境", "DEV / STG / PRD のいずれか (記入)"),
        ("テストアカウント", ""),
        ("実施者", ""),
        ("レビュー者", ""),
        ("承認者", ""),
    ]
    for i, (k, v) in enumerate(rows, start=6):
        _label(ws, f"B{i}", k)
        ws[f"C{i}"] = v
    _label(ws, "B15", "使い方")
    for i, line in enumerate([
        "・各カテゴリシートで「結果(OK/NG)」列にドロップダウンから選択。",
        "・NG の場合は「NG時の内容・備考」「エビデンスNo.」「実施日」を必ず記入。",
        "・エビデンス画像はエビデンスシートの該当ブロックに貼り付け (枠内に収める)。",
        "・1 シートにつき確認日・確認者をヘッダーに記入。",
    ], start=16):
        ws[f"B{i}"] = line
    _label(ws, "B21", "凡例")
    _label(ws, "B22", "記号")
    _label(ws, "C22", "意味")
    for i, (sym, mean) in enumerate([
        ("OK", "期待結果と一致"),
        ("NG", "期待結果と不一致 (要修正)"),
        ("-", "対象外 / スキップ"),
        ("保留", "判定保留 (要再確認)"),
    ], start=23):
        ws[f"B{i}"] = sym
        ws[f"C{i}"] = mean
    _label(ws, "B28", "改訂履歴")
    for j, h in enumerate(["バージョン", "日付", "変更内容", "担当"]):
        c = ws.cell(29, 2 + j, h)
        c.font = _BOLD
    ws.cell(30, 2, meta.get("version", "v1.0"))
    ws.cell(30, 3, meta.get("issued", ""))
    ws.cell(30, 4, "新規作成")
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    for col in "DEF":
        ws.column_dimensions[col].width = 16


def _build_category(ws, category, rows, meta):
    # header rows 1-2
    pairs1 = [("画面名", ""), ("機能カテゴリ", category), ("確認日", ""), ("確認者", "")]
    pairs2 = [("対象URL", meta.get("url", "")), ("環境", "DEV/STG"), ("アカウント/ロール", ""), ("バージョン", meta.get("version", "v1.0"))]
    for r, pairs in ((1, pairs1), (2, pairs2)):
        col = 1
        for k, v in pairs:
            kc = ws.cell(r, col, k)
            kc.font = _BOLD
            kc.fill = _HEADER_FILL
            ws.cell(r, col + 1, v)
            col += 2
    ws.cell(4, 1, f"【{category}】チェックリスト").font = _BOLD
    # column header row 6
    for j, h in enumerate(CATEGORY_COLUMNS, start=1):
        c = ws.cell(6, j, h)
        c.font = _BOLD
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = _WRAP
    # empty body rows
    first, last = 7, 7 + rows - 1
    for r in range(first, last + 1):
        for j in range(1, len(CATEGORY_COLUMNS) + 1):
            ws.cell(r, j).border = _BORDER
            ws.cell(r, j).alignment = _WRAP
    # OK/NG dropdown on the 結果 column (F)
    result_col = get_column_letter(CATEGORY_COLUMNS.index("結果(OK/NG)") + 1)
    dv = DataValidation(type="list", formula1=RESULT_CHOICES, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{result_col}{first}:{result_col}{last}")
    widths = {"A": 8, "B": 14, "C": 32, "D": 32, "E": 14, "F": 28, "G": 12, "H": 24, "I": 14, "J": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _build_evidence(ws, blocks):
    ws["B2"] = "エビデンス記録シート"
    ws["B2"].font = _TITLE_FONT
    ws["B3"] = "各チェック No. ごとに 1 ブロック。スクリーンショットは下部の枠内に貼り付けてください。"
    labels = ["エビデンスNo.", "関連チェックNo.", "取得日時", "実施環境", "操作手順詳細", "確認結果", "備考"]
    r = 5
    for _ in range(blocks):
        ws.cell(r, 2, "■ エビデンス No.").font = _BOLD
        r += 1
        for lab in labels:
            ws.cell(r, 2, lab).font = _BOLD
            ws.cell(r, 3).alignment = _WRAP
            r += 1
        ws.cell(r, 2, "スクリーンショット (この枠内に貼り付け)").font = _BOLD
        r += 1
        # tall merged paste frame B:H over 10 rows
        ws.merge_cells(start_row=r, start_column=2, end_row=r + 9, end_column=8)
        frame = ws.cell(r, 2)
        frame.border = _BORDER
        frame.alignment = Alignment(horizontal="center", vertical="center")
        for rr in range(r, r + 10):
            ws.row_dimensions[rr].height = 18
        r += 11  # frame + spacer
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40
    for col in "DEFGH":
        ws.column_dimensions[col].width = 14


def _build_data(ws, data_sets, empty_rows=6):
    ws["B2"] = "利用データ（テストデータ）一覧"
    ws["B2"].font = _TITLE_FONT
    ws["B3"] = "テストごとに利用データが異なる場合はここに定義し、各カテゴリ行の「利用データNo.」から参照する。"
    # header row 5
    for j, h in enumerate(DATA_COLUMNS, start=2):  # start at col B
        c = ws.cell(5, j, h)
        c.font = _BOLD
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = _WRAP
    # pre-filled defaults + empty rows for additions
    rows = list(data_sets) + [("", "", "", "", "")] * empty_rows
    for i, rec in enumerate(rows, start=6):
        for j, val in enumerate(rec, start=2):
            cell = ws.cell(i, j, val if val != "" else None)
            cell.border = _BORDER
            cell.alignment = _WRAP
    widths = {"B": 14, "C": 16, "D": 30, "E": 52, "F": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_template(path, categories=None, rows=12, evidence_blocks=6, meta=None):
    """Write an empty IT確認書 template workbook to `path`."""
    categories = categories or DEFAULT_CATEGORIES
    meta = meta or {}
    wb = Workbook()
    _build_cover(wb.active, meta)
    for i, cat in enumerate(categories, start=2):
        ws = wb.create_sheet(f"{i:02d}_{cat}")
        _build_category(ws, cat, rows, meta)
    _build_evidence(wb.create_sheet(f"{len(categories) + 2:02d}_エビデンス"), evidence_blocks)
    _build_data(
        wb.create_sheet(f"{len(categories) + 3:02d}_利用データ"),
        meta.get("data_sets") or DEFAULT_DATA_SETS,
        empty_rows=max(2, rows // 2),
    )
    wb.save(path)
    return [ws.title for ws in wb.worksheets]
