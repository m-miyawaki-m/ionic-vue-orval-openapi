"""Generate the evidence workbook: spec + measured results + manual screenshot cells."""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .results import join_status

EV_HEADERS = [
    "ID", "画面", "観点", "前提", "操作・入力", "期待結果",
    "実測", "実行日時", "スクショ貼付欄", "画像パス",
]

_STATUS_MAP = {"passed": "pass", "failed": "fail"}


def _status_label(item_id, results):
    r = join_status(item_id, results)
    if r is None:
        return "—"
    return _STATUS_MAP.get(r["status"], r["status"])


def _fill_sheet(ws, rows, results, run_time, screenshot_dir):
    ws.append(EV_HEADERS)
    shot_col = get_column_letter(EV_HEADERS.index("スクショ貼付欄") + 1)
    ws.column_dimensions[shot_col].width = 40  # wide cell for pasted screenshot
    for i, it in enumerate(rows, start=2):
        ws.append([
            it.id, it.screen, it.perspective, it.precondition, it.action,
            it.expected, _status_label(it.id, results), run_time,
            "",  # スクショ貼付欄: intentionally empty for manual paste
            f"{screenshot_dir}/{it.id}.png",
        ])
        ws.row_dimensions[i].height = 80  # tall row to hold a pasted image


def write_evidence_xlsx(items, results, run_time, path, screenshot_dir="out/screenshots"):
    """Write evidence.xlsx with 概要 / L1境界値 / 操作・状態遷移 sheets."""
    wb = Workbook()
    summary = wb.active
    summary.title = "概要"

    boundary_items = [it for it in items if it.perspective == "境界値"]
    other_items = [it for it in items if it.perspective != "境界値"]

    _fill_sheet(wb.create_sheet("L1境界値"), boundary_items, results, run_time, screenshot_dir)
    _fill_sheet(wb.create_sheet("操作・状態遷移"), other_items, results, run_time, screenshot_dir)

    total = len(items)
    passed = sum(1 for it in items if _status_label(it.id, results) == "pass")
    failed = sum(1 for it in items if _status_label(it.id, results) == "fail")
    summary.append(["生成日時", run_time])
    summary.append(["項目総数", total])
    summary.append(["実測 pass", passed])
    summary.append(["実測 fail", failed])
    summary.append(["未測(—)", total - passed - failed])
    summary.append([])
    summary.append(["注", "スクショ貼付欄は手動で画像を貼り付ける（画像パス列にファイル規約を記載済み）"])

    wb.save(path)
