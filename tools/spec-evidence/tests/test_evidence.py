import os
import tempfile
import unittest

from openpyxl import load_workbook

from spec_evidence.evidence import write_evidence_xlsx
from spec_evidence.model import TestItem

ITEMS = [
    TestItem("bnd:x", "item", "境界値", "", "name = 'a'", "有効(受理)"),
    TestItem("bnd:y", "item", "境界値", "", "name = ''", "無効(拒否)"),
    TestItem("op:1", "auth", "操作", "未認証", "login", "token保存"),
]
RESULTS = {
    "bnd:x": {"status": "passed", "duration": 1},
    "bnd:y": {"status": "failed", "duration": 1},
}


class TestEvidence(unittest.TestCase):
    def test_evidence(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "evidence.xlsx")
            write_evidence_xlsx(ITEMS, RESULTS, "2026-05-26T00:00:00", p)
            wb = load_workbook(p)
            self.assertIn("概要", wb.sheetnames)
            self.assertIn("L1境界値", wb.sheetnames)
            self.assertIn("操作・状態遷移", wb.sheetnames)

            ws = wb["L1境界値"]
            headers = [c.value for c in ws[1]]
            for h in ["実測", "実行日時", "スクショ貼付欄", "画像パス"]:
                self.assertIn(h, headers)
            self.assertEqual(ws.max_row, 3)  # header + 2 boundary items

            status_col = headers.index("実測") + 1
            self.assertEqual(ws.cell(row=2, column=status_col).value, "pass")
            self.assertEqual(ws.cell(row=3, column=status_col).value, "fail")

            other = wb["操作・状態遷移"]
            self.assertEqual(other.max_row, 2)  # header + 1 op item

            summ = wb["概要"]
            vals = {
                summ.cell(row=r, column=1).value: summ.cell(row=r, column=2).value
                for r in range(1, summ.max_row + 1)
            }
            self.assertEqual(vals["項目総数"], 3)
            self.assertEqual(vals["実測 pass"], 1)
            self.assertEqual(vals["実測 fail"], 1)
            self.assertEqual(vals["未測(—)"], 1)


if __name__ == "__main__":
    unittest.main()
