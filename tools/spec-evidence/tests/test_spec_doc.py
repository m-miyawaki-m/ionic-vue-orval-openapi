import os
import tempfile
import unittest

from openpyxl import load_workbook

from spec_evidence.model import TestItem
from spec_evidence.spec_doc import write_spec_md, write_spec_xlsx

ITEMS = [
    TestItem("id1", "item", "境界値", "", "name = 'a'", "有効(受理)"),
    TestItem("id2", "auth", "操作", "未認証", "login", "token保存"),
]


class TestSpecDoc(unittest.TestCase):
    def test_md(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "spec.md")
            text = write_spec_md(ITEMS, p)
            self.assertIn("| ID | 画面 | 観点 | 前提 | 操作・入力 | 期待結果 |", text)
            self.assertIn("id1", text)
            self.assertIn("token保存", text)
            self.assertTrue(os.path.exists(p))

    def test_xlsx(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "spec.xlsx")
            write_spec_xlsx(ITEMS, p)
            wb = load_workbook(p)
            ws = wb.active
            self.assertEqual(ws["A1"].value, "ID")
            self.assertEqual(ws.max_row, len(ITEMS) + 1)


if __name__ == "__main__":
    unittest.main()
