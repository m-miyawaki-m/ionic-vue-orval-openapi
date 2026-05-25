import os
import tempfile
import unittest

from openpyxl import load_workbook

from spec_evidence import cli

FIX = os.path.join(os.path.dirname(__file__), "fixtures")


class TestCli(unittest.TestCase):
    def test_cli_smoke(self):
        with tempfile.TemporaryDirectory() as d:
            rc = cli.main([
                "--boundary", os.path.join(FIX, "boundary.sample.json"),
                "--doc", os.path.join(FIX, "doc.sample.json"),
                "--results", os.path.join(FIX, "vitest.sample.json"),
                "--out-dir", d,
            ])
            self.assertEqual(rc, 0)
            for fn in ["test-spec.md", "test-spec.xlsx", "evidence.xlsx"]:
                self.assertTrue(os.path.exists(os.path.join(d, fn)), fn)
            wb = load_workbook(os.path.join(d, "evidence.xlsx"))
            self.assertIn("L1境界値", wb.sheetnames)


if __name__ == "__main__":
    unittest.main()
