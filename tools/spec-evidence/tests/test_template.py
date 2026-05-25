import os
import tempfile
import unittest

from openpyxl import load_workbook

from spec_evidence.template import CATEGORY_COLUMNS, DATA_COLUMNS, write_template


class TestTemplate(unittest.TestCase):
    def test_template_structure(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "tmpl.xlsx")
            sheets = write_template(p, categories=["画面表示", "API連携"], rows=5, evidence_blocks=2)
            # 01 cover / 02-03 categories / 04 evidence / 05 data
            self.assertEqual(sheets, ["01_表紙", "02_画面表示", "03_API連携", "04_エビデンス", "05_利用データ"])

            wb = load_workbook(p)

            # cover: title + legend; NO 利用データ section on the cover
            cover = wb["01_表紙"]
            self.assertEqual(cover["B2"].value, "IT テスト確認書")
            self.assertEqual([cover.cell(r, 2).value for r in range(23, 27)], ["OK", "NG", "-", "保留"])
            cover_labels = [cover.cell(r, 2).value for r in range(1, cover.max_row + 1)]
            self.assertNotIn("利用データ（テストデータ）", cover_labels)

            # category sheet: 10-column header incl. 利用データNo.; OK/NG dropdown present
            cat = wb["02_画面表示"]
            headers = [cat.cell(6, c).value for c in range(1, len(CATEGORY_COLUMNS) + 1)]
            self.assertEqual(headers, CATEGORY_COLUMNS)
            self.assertIn("利用データNo.", headers)
            self.assertTrue(len(cat.data_validations.dataValidation) >= 1)

            # evidence sheet: 2 block markers
            ev = wb["04_エビデンス"]
            markers = [c.value for row in ev.iter_rows() for c in row if c.value == "■ エビデンス No."]
            self.assertEqual(len(markers), 2)

            # data sheet: header + at least the default DT-001 row
            data = wb["05_利用データ"]
            data_headers = [data.cell(5, c).value for c in range(2, 2 + len(DATA_COLUMNS))]
            self.assertEqual(data_headers, DATA_COLUMNS)
            ids = [data.cell(r, 2).value for r in range(6, data.max_row + 1)]
            self.assertIn("DT-001", ids)


if __name__ == "__main__":
    unittest.main()
